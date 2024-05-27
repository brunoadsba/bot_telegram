import telebot
from telebot import types
from datetime import datetime, timedelta
import threading
from openpyxl import Workbook, load_workbook
import os
import time

# Chave API do bot
chave_API = "7103427784:AAHmjR-l_ZGdwW1sJwlCu5VGx7ow8D2Bwp4"

# Cria a instância do bot
bot = telebot.TeleBot(chave_API)

# Variável global para armazenar o estado da conversa
estados_conversa = {}

# Função para enviar as opções de escolha
def enviar_opcoes(chat_id):
    texto_opcoes = """
    **Escolha uma opção:**

    /inserir Inserir dados na ficha de EPI
    
    /anotar Anotar informação importante
    """
    bot.send_message(chat_id, texto_opcoes)

# Função para remover o teclado da tela
def remover_teclado(chat_id):
    remove_teclado = types.ReplyKeyboardRemove()
    bot.send_message(chat_id, "Opções removidas.", reply_markup=remove_teclado)

# Função para verificar se o usuário está ocioso há mais de 1 minuto
def verificar_inatividade():
    for chat_id, estado in list(estados_conversa.items()):
        ultima_interacao = estado.get("ultima_interacao")
        if ultima_interacao:
            diferenca = datetime.now() - ultima_interacao
            if diferenca > timedelta(minutes=1):
                # Se não houver interação por 1 minuto, encerra a conversa
                bot.send_message(chat_id, "Se não tem mais informações, eu irei finalizar o chat.\nFique à vontade para me chamar a qualquer momento!")
                remover_teclado(chat_id)  # Remover teclado
                estados_conversa.pop(chat_id)
    # Agenda a próxima verificação após 60 segundos
    threading.Timer(60, verificar_inatividade).start()

# Inicia a verificação de inatividade
verificar_inatividade()

# Função para inserir os dados na planilha
def inserir_na_planilha(data_entrega, nome, funcao, epi, observacao):
    arquivo_planilha = "planilha_epi.xlsx"
    
    # Verifica se o arquivo da planilha existe
    if os.path.isfile(arquivo_planilha):
        # Espera até que o arquivo esteja disponível para escrita
        while True:
            try:
                workbook = load_workbook(arquivo_planilha)
                sheet = workbook.active
                break
            except PermissionError:
                time.sleep(1)  # Espera 1 segundo antes de tentar novamente
        
        # Insere os dados na linha abaixo da última linha preenchida
        nova_linha = [data_entrega, nome, funcao, epi, observacao]
        sheet.append(nova_linha)
        workbook.save(arquivo_planilha)
    else:
        # Se o arquivo não existir, cria uma nova planilha
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Data de Entrega", "Nome", "Função", "EPI", "Observação"])  # Cabeçalho da planilha
            nova_linha = [data_entrega, nome, funcao, epi, observacao]
            sheet.append(nova_linha)
            workbook.save(arquivo_planilha)
        except Exception as e:
            print("Erro ao criar ou abrir a planilha:", e)
            return False

    return True

# Handler para o comando /inserir
@bot.message_handler(commands=['inserir'])
def inserir(mensagem):
    # Define o estado da conversa como "aguardando_nome"
    estados_conversa[mensagem.chat.id] = {"estado": "aguardando_nome", "ultima_interacao": datetime.now()}
    bot.send_message(mensagem.chat.id, "Diga o nome completo do funcionário para inserir dados na ficha de EPI:")

# Handler para processar o nome completo do funcionário
@bot.message_handler(func=lambda mensagem: estados_conversa.get(mensagem.chat.id, {}).get("estado") == "aguardando_nome")
def processar_nome_completo(mensagem):
    # Armazena o nome completo do funcionário
    estados_conversa[mensagem.chat.id]["nome_completo"] = mensagem.text
    estados_conversa[mensagem.chat.id]["ultima_interacao"] = datetime.now()
    
    # Define o estado da conversa como "aguardando_funcao"
    estados_conversa[mensagem.chat.id]["estado"] = "aguardando_funcao"

    # Pergunta a função do colaborador
    bot.send_message(mensagem.chat.id, "Qual a função do colaborador?")

# Handler para processar a função do colaborador
@bot.message_handler(func=lambda mensagem: estados_conversa.get(mensagem.chat.id, {}).get("estado") == "aguardando_funcao")
def processar_funcao(mensagem):
    # Armazena a função do colaborador
    estados_conversa[mensagem.chat.id]["funcao"] = mensagem.text
    estados_conversa[mensagem.chat.id]["ultima_interacao"] = datetime.now()
    
    # Define o estado da conversa como "aguardando_data_entrega_epi"
    estados_conversa[mensagem.chat.id]["estado"] = "aguardando_data_entrega_epi"

    # Solicita a data de entrega do EPI
    bot.send_message(mensagem.chat.id, "Informe a data de entrega do EPI (no formato DD/MM/AAAA) para a nova ficha de EPI:")

# Handler para processar a data de entrega do EPI
@bot.message_handler(func=lambda mensagem: estados_conversa.get(mensagem.chat.id, {}).get("estado") == "aguardando_data_entrega_epi")
def processar_data_entrega_epi(mensagem):
    # Verifica se a data de entrega está no formato correto
    if len(mensagem.text) != 10 or not mensagem.text[0:2].isdigit() or not mensagem.text[3:5].isdigit() or not mensagem.text[6:].isdigit() or mensagem.text[2] != '/' or mensagem.text[5] != '/':
        bot.send_message(mensagem.chat.id, "Formato de data inválido. Por favor, informe a data de entrega no formato DD/MM/AAAA.")
        return

    # Armazena a data de entrega do EPI
    estados_conversa[mensagem.chat.id]["data_entrega_epi"] = mensagem.text
    estados_conversa[mensagem.chat.id]["ultima_interacao"] = datetime.now()

    # Define o estado da conversa como "aguardando_tipo_epi"
    estados_conversa[mensagem.chat.id]["estado"] = "aguardando_tipo_epi"

    # Solicita o tipo de EPI
    bot.send_message(mensagem.chat.id, "Informe o tipo de EPI para a nova ficha de EPI:")

# Handler para processar o tipo de EPI
@bot.message_handler(func=lambda mensagem: estados_conversa.get(mensagem.chat.id, {}).get("estado") == "aguardando_tipo_epi")
def processar_tipo_epi(mensagem):
    # Armazena o tipo de EPI
    tipo_epi = mensagem.text

    # Recupera os dados armazenados
    dados = estados_conversa.get(mensagem.chat.id, {})
    nome_completo = dados.get("nome_completo")
    funcao = dados.get("funcao")
    data_entrega_epi = dados.get("data_entrega_epi")

    # Pergunta se deseja inserir alguma observação
    bot.send_message(mensagem.chat.id, "Deseja adicionar alguma observação? (Digite a observação ou envie /pular para continuar)")

    # Define o estado da conversa como "aguardando_observacao"
    estados_conversa[mensagem.chat.id]["estado"] = "aguardando_observacao"

    # Armazena os dados necessários para inserir na planilha
    estados_conversa[mensagem.chat.id]["tipo_epi"] = tipo_epi
    estados_conversa[mensagem.chat.id]["nome_completo"] = nome_completo
    estados_conversa[mensagem.chat.id]["funcao"] = funcao
    estados_conversa[mensagem.chat.id]["data_entrega_epi"] = data_entrega_epi

# Handler para processar a observação
@bot.message_handler(func=lambda mensagem: estados_conversa.get(mensagem.chat.id, {}).get("estado") == "aguardando_observacao")
def processar_observacao(mensagem):
    # Armazena a observação, se fornecida
    observacao = mensagem.text if mensagem.text != "/pular" else ""
    
    # Recupera os dados armazenados
    dados = estados_conversa.get(mensagem.chat.id, {})
    tipo_epi = dados.get("tipo_epi")
    nome_completo = dados.get("nome_completo")
    funcao = dados.get("funcao")
    data_entrega_epi = dados.get("data_entrega_epi")

    # Insere os dados na planilha
    sucesso = inserir_na_planilha(data_entrega_epi, nome_completo, funcao, tipo_epi, observacao)

    if sucesso:
        # Informa ao usuário que a inserção foi concluída
        bot.send_message(mensagem.chat.id, "Dados inseridos com sucesso!")
        estados_conversa[mensagem.chat.id]["ultima_interacao"] = datetime.now()

        # Pergunta se deseja continuar ou finalizar
        teclado = types.ReplyKeyboardMarkup(resize_keyboard=True)
        teclado.add(types.KeyboardButton("Continuar"), types.KeyboardButton("Finalizar"))
        bot.send_message(mensagem.chat.id, "Deseja continuar?", reply_markup=teclado)

        # Define o estado da conversa como "aguardando_opcao_continuar"
        estados_conversa[mensagem.chat.id]["estado"] = "aguardando_opcao_continuar"
    else:
        # Informa ao usuário que ocorreu um erro ao inserir os dados na planilha
        bot.send_message(mensagem.chat.id, "Ocorreu um erro ao inserir os dados na planilha. Por favor, tente novamente mais tarde.")

# Handler para processar a opção após a inserção dos dados
@bot.message_handler(func=lambda mensagem: estados_conversa.get(mensagem.chat.id, {}).get("estado") == "aguardando_opcao_continuar")
def processar_opcao_continuar(mensagem):
    # Se a opção for "Continuar", volta para as opções padrão
    if mensagem.text == "Continuar":
        enviar_opcoes(mensagem.chat.id)
    # Se a opção for "Finalizar", remove o teclado
    elif mensagem.text == "Finalizar":
        bot.send_message(mensagem.chat.id, "Opção selecionada: Finalizar.")
        remover_teclado(mensagem.chat.id)  # Remover teclado
        estados_conversa.pop(mensagem.chat.id)

# Handler para o comando /criar
@bot.message_handler(commands=['criar'])
def criar(mensagem):
    # Define o estado da conversa como "aguardando_nome"
    estados_conversa[mensagem.chat.id] = {"estado": "aguardando_nome", "ultima_interacao": datetime.now()}
    bot.send_message(mensagem.chat.id, "Diga o nome completo do funcionário para criar a nova ficha de EPI:")

# Handler para processar o nome completo do funcionário para criar uma nova ficha de EPI
@bot.message_handler(func=lambda mensagem: estados_conversa.get(mensagem.chat.id, {}).get("estado") == "aguardando_nome")
def processar_nome_completo_criar(mensagem):
    # Armazena o nome completo do funcionário
    estados_conversa[mensagem.chat.id]["nome_completo"] = mensagem.text
    estados_conversa[mensagem.chat.id]["ultima_interacao"] = datetime.now()
    
    # Define o estado da conversa como "aguardando_funcao"
    estados_conversa[mensagem.chat.id]["estado"] = "aguardando_funcao"

    # Pergunta a função do colaborador
    bot.send_message(mensagem.chat.id, "Qual a função do colaborador para a nova ficha de EPI?")

# Handler para processar a função do colaborador para criar uma nova ficha de EPI
@bot.message_handler(func=lambda mensagem: estados_conversa.get(mensagem.chat.id, {}).get("estado") == "aguardando_funcao")
def processar_funcao_criar(mensagem):
    # Armazena a função do colaborador
    estados_conversa[mensagem.chat.id]["funcao"] = mensagem.text
    estados_conversa[mensagem.chat.id]["ultima_interacao"] = datetime.now()
    
    # Define o estado da conversa como "aguardando_data_entrega_epi"
    estados_conversa[mensagem.chat.id]["estado"] = "aguardando_data_entrega_epi"

    # Solicita a data de entrega do EPI
    bot.send_message(mensagem.chat.id, "Informe a data de entrega do EPI (no formato DD/MM/AAAA) para a nova ficha de EPI:")

# Handler para processar a anotação do usuário
@bot.message_handler(commands=['anotar'])
def anotar(mensagem):
    # Define o estado da conversa como "aguardando_anotacao"
    estados_conversa[mensagem.chat.id] = {"estado": "aguardando_anotacao", "ultima_interacao": datetime.now()}
    bot.send_message(mensagem.chat.id, "O que gostaria de anotar?")

# Handler para processar a anotação do usuário
@bot.message_handler(func=lambda mensagem: estados_conversa.get(mensagem.chat.id, {}).get("estado") == "aguardando_anotacao")
def processar_anotacao(mensagem):
    # Armazena a anotação do usuário
    anotacao = mensagem.text
    estados_conversa[mensagem.chat.id]["ultima_interacao"] = datetime.now()
    
    # Realiza o que desejar com a anotação, como salvar em um banco de dados, por exemplo
    print(f"Anotação: {anotacao}")

    # Informa ao usuário que a anotação foi realizada com sucesso
    bot.send_message(mensagem.chat.id, "Anotação realizada com sucesso!")

    # Pergunta se deseja continuar ou finalizar
    teclado = types.ReplyKeyboardMarkup(resize_keyboard=True)
    teclado.add(types.KeyboardButton("Continuar"), types.KeyboardButton("Finalizar"))
    bot.send_message(mensagem.chat.id, "Deseja continuar?", reply_markup=teclado)

    # Define o estado da conversa como "aguardando_opcao_continuar"
    estados_conversa[mensagem.chat.id]["estado"] = "aguardando_opcao_continuar"

# Handler para processar o comando "menu inicial"
@bot.message_handler(func=lambda mensagem: mensagem.text.lower() == 'menu inicial')
def menu_inicial(mensagem):
    estados_conversa.pop(mensagem.chat.id, None)  # Limpa o estado da conversa
    enviar_opcoes(mensagem.chat.id)

# Handler padrão para responder a outras mensagens
@bot.message_handler(func=lambda mensagem: True)
def verificar(mensagem):
    # Se o usuário enviar uma mensagem inesperada, informa as opções disponíveis
    enviar_opcoes(mensagem.chat.id)

bot.polling()



