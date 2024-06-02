import telebot
from telebot import types
from datetime import datetime, timedelta
import threading
from openpyxl import Workbook, load_workbook
import os
import time

# Chave API do bot
API_KEY = " "

# Cria a instância do bot
bot = telebot.TeleBot(API_KEY)

# Gerenciamento de estado da conversa
class EstadoConversa:
    def __init__(self):
        self.estados = {}

    def get_estado(self, chat_id):
        return self.estados.get(chat_id, {})

    def set_estado(self, chat_id, estado):
        self.estados[chat_id] = estado

    def remove_estado(self, chat_id):
        if chat_id in self.estados:
            del self.estados[chat_id]

estado_conversa = EstadoConversa()

# Função para enviar as opções de escolha
def enviar_opcoes(chat_id):
    texto_opcoes = """
    **Escolha uma opção:**

    /inserir Inserir dados na ficha de EPI
    
    /anotar Anotar informação importante
    """
    bot.send_message(chat_id, texto_opcoes)

# Função para remover o teclado da tela
def remover_teclado(chat_id):
    remove_teclado = types.ReplyKeyboardRemove()
    bot.send_message(chat_id, "Opções removidas.", reply_markup=remove_teclado)

# Função para verificar se o usuário está ocioso há mais de 1 minuto
def verificar_inatividade():
    agora = datetime.now()
    for chat_id, estado in list(estado_conversa.estados.items()):
        ultima_interacao = estado.get("ultima_interacao")
        if ultima_interacao and (agora - ultima_interacao > timedelta(minutes=1)):
            bot.send_message(chat_id, "Se não tem mais informações, eu irei finalizar o chat.\nFique à vontade para me chamar a qualquer momento!")
            remover_teclado(chat_id)
            estado_conversa.remove_estado(chat_id)
    threading.Timer(60, verificar_inatividade).start()

# Inicia a verificação de inatividade
verificar_inatividade()

# Função para inserir os dados na planilha
def inserir_na_planilha(data_entrega, nome, funcao, epi, observacao):
    arquivo_planilha = "planilha_epi.xlsx"
    
    try:
        if os.path.isfile(arquivo_planilha):
            while True:
                try:
                    workbook = load_workbook(arquivo_planilha)
                    break
                except PermissionError:
                    time.sleep(1)
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Data de Entrega", "Nome", "Função", "EPI", "Observação"])
            workbook.save(arquivo_planilha)
        
        sheet = workbook.active
        sheet.append([data_entrega, nome, funcao, epi, observacao])
        workbook.save(arquivo_planilha)
    except Exception as e:
        print(f"Erro ao manipular a planilha: {e}")
        return False

    return True

# Handler para o comando /inserir
@bot.message_handler(commands=['inserir'])
def inserir(mensagem):
    estado_conversa.set_estado(mensagem.chat.id, {"estado": "aguardando_nome", "ultima_interacao": datetime.now()})
    bot.send_message(mensagem.chat.id, "Diga o nome completo do funcionário para inserir dados na ficha de EPI:")

# Handlers para processar os diferentes estados da conversa
@bot.message_handler(func=lambda mensagem: estado_conversa.get_estado(mensagem.chat.id).get("estado") == "aguardando_nome")
def processar_nome_completo(mensagem):
    estado = estado_conversa.get_estado(mensagem.chat.id)
    estado["nome_completo"] = mensagem.text
    estado["estado"] = "aguardando_funcao"
    estado["ultima_interacao"] = datetime.now()
    bot.send_message(mensagem.chat.id, "Qual a função do colaborador?")

@bot.message_handler(func=lambda mensagem: estado_conversa.get_estado(mensagem.chat.id).get("estado") == "aguardando_funcao")
def processar_funcao(mensagem):
    estado = estado_conversa.get_estado(mensagem.chat.id)
    estado["funcao"] = mensagem.text
    estado["estado"] = "aguardando_data_entrega_epi"
    estado["ultima_interacao"] = datetime.now()
    bot.send_message(mensagem.chat.id, "Informe a data de entrega do EPI (no formato DD/MM/AAAA):")

@bot.message_handler(func=lambda mensagem: estado_conversa.get_estado(mensagem.chat.id).get("estado") == "aguardando_data_entrega_epi")
def processar_data_entrega_epi(mensagem):
    if len(mensagem.text) != 10 or not all(x.isdigit() for x in mensagem.text.replace('/', '')):
        bot.send_message(mensagem.chat.id, "Formato de data inválido. Por favor, informe a data de entrega no formato DD/MM/AAAA.")
        return

    estado = estado_conversa.get_estado(mensagem.chat.id)
    estado["data_entrega_epi"] = mensagem.text
    estado["estado"] = "aguardando_tipo_epi"
    estado["ultima_interacao"] = datetime.now()
    bot.send_message(mensagem.chat.id, "Informe o tipo de EPI:")

@bot.message_handler(func=lambda mensagem: estado_conversa.get_estado(mensagem.chat.id).get("estado") == "aguardando_tipo_epi")
def processar_tipo_epi(mensagem):
    estado = estado_conversa.get_estado(mensagem.chat.id)
    estado["tipo_epi"] = mensagem.text
    estado["estado"] = "aguardando_observacao"
    estado["ultima_interacao"] = datetime.now()
    bot.send_message(mensagem.chat.id, "Deseja adicionar alguma observação? (Digite a observação ou envie /pular para continuar)")

@bot.message_handler(func=lambda mensagem: estado_conversa.get_estado(mensagem.chat.id).get("estado") == "aguardando_observacao")
def processar_observacao(mensagem):
    observacao = mensagem.text if mensagem.text != "/pular" else ""
    estado = estado_conversa.get_estado(mensagem.chat.id)
    
    sucesso = inserir_na_planilha(estado["data_entrega_epi"], estado["nome_completo"], estado["funcao"], estado["tipo_epi"], observacao)
    
    if sucesso:
        bot.send_message(mensagem.chat.id, "Dados inseridos com sucesso!")
        teclado = types.ReplyKeyboardMarkup(resize_keyboard=True)
        teclado.add(types.KeyboardButton("Continuar"), types.KeyboardButton("Finalizar"))
        bot.send_message(mensagem.chat.id, "Deseja continuar?", reply_markup=teclado)
        estado["estado"] = "aguardando_opcao_continuar"
    else:
        bot.send_message(mensagem.chat.id, "Ocorreu um erro ao inserir os dados na planilha. Por favor, tente novamente mais tarde.")

@bot.message_handler(func=lambda mensagem: estado_conversa.get_estado(mensagem.chat.id).get("estado") == "aguardando_opcao_continuar")
def processar_opcao_continuar(mensagem):
    if mensagem.text == "Continuar":
        enviar_opcoes(mensagem.chat.id)
    elif mensagem.text == "Finalizar":
        bot.send_message(mensagem.chat.id, "Opção selecionada: Finalizar.")
        remover_teclado(mensagem.chat.id)
        estado_conversa.remove_estado(mensagem.chat.id)

# Handler para o comando /anotar
@bot.message_handler(commands=['anotar'])
def anotar(mensagem):
    estado_conversa.set_estado(mensagem.chat.id, {"estado": "aguardando_anotacao", "ultima_interacao": datetime.now()})
    bot.send_message(mensagem.chat.id, "O que gostaria de anotar?")

@bot.message_handler(func=lambda mensagem: estado_conversa.get_estado(mensagem.chat.id).get("estado") == "aguardando_anotacao")
def processar_anotacao(mensagem):
    anotacao = mensagem.text
    print(f"Anotação: {anotacao}")
    bot.send_message(mensagem.chat.id, "Anotação realizada com sucesso!")
    
    teclado = types.ReplyKeyboardMarkup(resize_keyboard=True)
    teclado.add(types.KeyboardButton("Continuar"), types.KeyboardButton("Finalizar"))
    bot.send_message(mensagem.chat.id, "Deseja continuar?", reply_markup=teclado)
    estado_conversa.get_estado(mensagem.chat.id)["estado"] = "aguardando_opcao_continuar"

# Handler para o comando "menu inicial"
@bot.message_handler(func=lambda mensagem: mensagem.text.lower() == 'menu inicial')
def menu_inicial(mensagem):
    estado_conversa.remove_estado(mensagem.chat.id)
    enviar_opcoes(mensagem.chat.id)

# Handler padrão para responder a outras mensagens
@bot.message_handler(func=lambda mensagem: True)
def verificar(mensagem):
    enviar_opcoes(mensagem.chat.id)

bot.polling()
